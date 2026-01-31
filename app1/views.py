from django.shortcuts import render, redirect
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from django.contrib.auth import logout
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_POST
from django.db import models
import json
from .models import ProductoAlmacen, Unidad,MovimientoInventario, PedidoCompra

def inicio(request):
    return render(request, 'home.html')

def almacenes(request):
    return render(request, 'almacenes.html')

def InventrioAG(request):
    productos = ProductoAlmacen.objects.filter(ubicacion_almacen='AG').order_by('-fecha_ingreso')
    
    # DEBUG: Ver en consola cuántos productos hay
    print(f"Productos encontrados: {productos.count()}")
    for p in productos:
        print(f"- {p.nombre} ({p.ubicacion_almacen})")
    
    context = {
        'productos': productos,
        'total_productos': productos.count()
    }
    return render(request, 'almacenes/almgeneral/InventarioAG.html', context)

def agregar(request):
    # Esta vista está mal - deberías usar agregar_producto
    unidades = Unidad.objects.all().order_by('nombre')
    return render(request, 'almacenes/almgeneral/agregar.html', {'unidades': unidades})

def perfil(request):
    return render(request, 'perfil.html')

def configuracion(request):
    return render(request, 'configuracion.html')

def logout_view(request):
    logout(request)
    return redirect('inicio')

@require_POST
def crear_unidad(request):
    try:
        data = json.loads(request.body)
        nombre = data.get('nombre', '').strip()
        abreviatura = data.get('abreviatura', '').strip()
        
        if not nombre:
            return JsonResponse({'success': False, 'error': 'El nombre es requerido'})
        
        # Verificar si ya existe
        if Unidad.objects.filter(nombre__iexact=nombre).exists():
            return JsonResponse({'success': False, 'error': 'Esta unidad ya existe'})
        
        unidad = Unidad.objects.create(
            nombre=nombre,
            abreviatura=abreviatura,
            activo=True  # AGREGA ESTO
        )
        
        return JsonResponse({
            'success': True,
            'unidad': {
                'id': unidad.id,
                'nombre': unidad.nombre,
                'abreviatura': unidad.abreviatura
            }
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

def agregar_producto(request):
    if request.method == 'POST':
        try:
            # DEBUG: Ver qué datos llegan
            print("=== DATOS RECIBIDOS ===")
            print(f"Nombre: {request.POST.get('nombre')}")
            print(f"Descripción: {request.POST.get('descripcion')}")
            print(f"Ubicación: {request.POST.get('ubicacion_almacen')}")
            print(f"Estante: {request.POST.get('estante')}")
            print(f"Cantidad: {request.POST.get('cantidad')}")
            print(f"Unidad: {request.POST.get('unidad')}")
            print(f"Estado: {request.POST.get('estado')}")
            
            producto = ProductoAlmacen.objects.create(
                nombre=request.POST.get('nombre'),
                descripcion=request.POST.get('descripcion'),
                ubicacion_almacen=request.POST.get('ubicacion_almacen'),
                estante=request.POST.get('estante'),
                cantidad=request.POST.get('cantidad'),
                unidad=request.POST.get('unidad'),
                estado=request.POST.get('estado'),
                observaciones=request.POST.get('observaciones', '')
            )
            
            print(f"✓ Producto creado: {producto.id_producto} - {producto.nombre}")
            messages.success(request, f'Producto "{producto.nombre}" agregado exitosamente')
            return redirect('InventarioAG')  # ← IMPORTANTE: debe coincidir con el name en urls.py
            
        except Exception as e:
            print(f"✗ ERROR: {str(e)}")
            messages.error(request, f'Error al agregar producto: {str(e)}')
    
    unidades = Unidad.objects.all().order_by('nombre')
    context = {
        'unidades': unidades
    }
    return render(request, 'almacenes/almgeneral/agregar_producto.html', context)
def importar_excel(request):
    if request.method == 'POST' and request.FILES.get('archivo_excel'):
        try:
            archivo = request.FILES['archivo_excel']
            
            # Leer el Excel
            df = pd.read_excel(archivo)
            
            # Validar columnas OBLIGATORIAS
            columnas_requeridas = ['codigo_almacen', 'codigo_producto', 'nombre', 'cantidad', 'unidad', 'estado']
            columnas_faltantes = [col for col in columnas_requeridas if col not in df.columns]
            
            if columnas_faltantes:
                messages.error(request, f'Faltan columnas obligatorias: {", ".join(columnas_faltantes)}')
                return redirect('importar_excel')
            
            productos_creados = 0
            productos_actualizados = 0
            unidades_creadas = 0
            errores = []
            
            # Mapeo de códigos de almacén a ubicación
            codigo_a_ubicacion = {
                '01': 'AG',
                '02': 'AD',
                '03': 'IU',
            }
            
            for index, row in df.iterrows():
                try:
                    # Validar código de almacén
                    codigo_almacen = str(row['codigo_almacen']).strip()
                    if codigo_almacen not in codigo_a_ubicacion:
                        errores.append(f"Fila {index+2}: Código de almacén '{codigo_almacen}' no válido (debe ser 01, 02 o 03)")
                        continue
                    
                    ubicacion = codigo_a_ubicacion[codigo_almacen]
                    
                    # Validar código de producto
                    codigo_producto = str(row['codigo_producto']).strip()
                    if not codigo_producto or codigo_producto == 'nan':
                        errores.append(f"Fila {index+2}: El código de producto es obligatorio")
                        continue
                    
                    # Validar nombre
                    nombre = str(row['nombre']).strip()
                    if not nombre or nombre == 'nan':
                        errores.append(f"Fila {index+2}: El nombre es obligatorio")
                        continue
                    
                    # Validar cantidad
                    try:
                        cantidad = int(row['cantidad'])
                        if cantidad < 0:
                            errores.append(f"Fila {index+2}: La cantidad no puede ser negativa")
                            continue
                    except:
                        errores.append(f"Fila {index+2}: Cantidad inválida")
                        continue
                    
                    # PROCESAR UNIDAD DE MEDIDA
                    unidad_str = str(row['unidad']).strip()
                    if not unidad_str or unidad_str == 'NAN':
                        errores.append(f"Fila {index+2}: La unidad es obligatoria")
                        continue
                    
                    # Buscar o crear la unidad (auto-crear si no existe)
                    unidad_obj = Unidad.objects.filter(
                        models.Q(nombre__iexact=unidad_str) |
                        models.Q(abreviatura__iexact=unidad_str)
                    ).first()
                    
                    if not unidad_obj:
                        # Crear nueva unidad automáticamente
                        unidad_obj = Unidad.objects.create(
                            nombre=unidad_str,
                            abreviatura=unidad_str[:10] if len(unidad_str) <= 10 else '',
                            activo=True
                        )
                        unidades_creadas += 1
                    
                    # Validar estado
                    estado = str(row['estado']).upper().strip()
                    if estado not in ['DISP', 'AGOT', 'BAJO']:
                        errores.append(f"Fila {index+2}: Estado '{estado}' no válido (debe ser DISP, AGOT o BAJO)")
                        continue
                    
                    # Campos opcionales
                    descripcion = str(row.get('descripcion', '')).strip()
                    if descripcion == 'nan':
                        descripcion = ''
                    
                    estante = str(row.get('estante', '')).strip()
                    if estante == 'nan':
                        estante = ''
                    
                    observaciones = str(row.get('observaciones', '')).strip()
                    if observaciones == 'nan':
                        observaciones = ''
                    
                    # BUSCAR SI EL PRODUCTO YA EXISTE
                    producto_existente = ProductoAlmacen.objects.filter(codigo_producto=codigo_producto).first()
                    
                    if producto_existente:
                        # ACTUALIZAR producto existente
                        cantidad_anterior = producto_existente.cantidad
                        estante_anterior = producto_existente.estante
                        
                        # Sumar la cantidad
                        producto_existente.cantidad += cantidad
                        
                        # Actualizar estante si cambió
                        if estante and estante != producto_existente.estante:
                            producto_existente.estante = estante
                        
                        # Actualizar unidad si cambió
                        if producto_existente.unidad != unidad_obj:
                            producto_existente.unidad = unidad_obj
                        
                        # Actualizar otros campos
                        producto_existente.estado = estado
                        if observaciones:
                            if producto_existente.observaciones:
                                producto_existente.observaciones += f"\n[{pd.Timestamp.now().strftime('%Y-%m-%d %H:%M')}] {observaciones}"
                            else:
                                producto_existente.observaciones = observaciones
                        
                        producto_existente.save()
                        
                        # Registrar el movimiento en el historial
                        MovimientoInventario.objects.create(
                            producto=producto_existente,
                            tipo_movimiento='ENTRADA',
                            cantidad=cantidad,
                            cantidad_anterior=cantidad_anterior,
                            cantidad_nueva=producto_existente.cantidad,
                            estante_anterior=estante_anterior,
                            estante_nuevo=producto_existente.estante,
                            observacion=f"Importación Excel: {observaciones}" if observaciones else "Importación Excel",
                            usuario=request.user.username if request.user.is_authenticated else 'Sistema'
                        )
                        
                        productos_actualizados += 1
                        
                    else:
                        # CREAR nuevo producto
                        nuevo_producto = ProductoAlmacen.objects.create(
                            codigo_producto=codigo_producto,
                            nombre=nombre,
                            descripcion=descripcion,
                            ubicacion_almacen=ubicacion,
                            estante=estante if estante else None,
                            cantidad=cantidad,
                            unidad=unidad_obj,
                            estado=estado,
                            observaciones=observaciones if observaciones else None
                        )
                        
                        # Registrar el movimiento inicial
                        MovimientoInventario.objects.create(
                            producto=nuevo_producto,
                            tipo_movimiento='ENTRADA',
                            cantidad=cantidad,
                            cantidad_anterior=0,
                            cantidad_nueva=cantidad,
                            estante_anterior=None,
                            estante_nuevo=estante if estante else None,
                            observacion=f"Creación inicial: {observaciones}" if observaciones else "Creación inicial",
                            usuario=request.user.username if request.user.is_authenticated else 'Sistema'
                        )
                        
                        productos_creados += 1
                    
                except Exception as e:
                    errores.append(f"Fila {index+2}: {str(e)}")
            
            # Mensajes de resultado
            if productos_creados > 0:
                messages.success(request, f'✅ Se crearon {productos_creados} productos nuevos')
            
            if productos_actualizados > 0:
                messages.success(request, f'✅ Se actualizaron {productos_actualizados} productos existentes')
            
            if unidades_creadas > 0:
                messages.info(request, f'ℹ️ Se crearon {unidades_creadas} nuevas unidades de medida')
            
            if errores:
                messages.warning(request, f'⚠️ Se encontraron {len(errores)} errores')
                for error in errores[:10]:
                    messages.error(request, error)
            
            if productos_creados > 0 or productos_actualizados > 0:
                return redirect('InventarioAG')
            else:
                return redirect('importar_excel')
            
        except Exception as e:
            messages.error(request, f'❌ Error al procesar el archivo: {str(e)}')
            return redirect('importar_excel')
    
    # GET: Mostrar formulario con unidades disponibles
    unidades_disponibles = Unidad.objects.filter(activo=True).order_by('nombre')
    
    return render(request, 'almacenes/importar_excel.html', {
        'unidades_disponibles': unidades_disponibles
    })
def descargar_plantilla(request):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from django.http import HttpResponse
    
    # Crear workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Plantilla Productos"
    
    # Headers - ORDEN CORRECTO
    headers = ['codigo_almacen', 'codigo_producto', 'nombre', 'descripcion', 'cantidad', 'unidad', 'estado', 'estante', 'observaciones']
    
    # Estilo de headers
    header_fill = PatternFill(start_color="148129", end_color="148129", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    
    # Fila de ayuda (comentarios)
    ayudas = [
        '01, 02 o 03',
        'Ej: PRD-001',
        'Obligatorio',
        'Opcional',
        'Obligatorio (número)',
        'Ej: Unidad, Caja, Kg',
        'DISP, AGOT o BAJO',
        'Opcional (Ej: A1, B2)',
        'Opcional'
    ]
    
    for col, ayuda in enumerate(ayudas, 1):
        cell = ws.cell(row=2, column=col, value=ayuda)
        cell.font = Font(color="666666", italic=True, size=9)
        cell.alignment = Alignment(horizontal="center")
    
    # Datos de ejemplo
    ejemplos = [
        ['01', 'PRD-001', 'Balón de Fútbol', 'Balón profesional N°5', 50, 'Unidad', 'DISP', 'A1', 'Stock nuevo'],
        ['01', 'PRD-002', 'Papel Bond A4', 'Paquete de 500 hojas', 100, 'Paquete', 'DISP', 'B2', ''],
        ['02', 'PRD-003', 'Raqueta Tenis', '', 15, 'Unidad', 'BAJO', 'C1', 'Reabastecer'],
        ['03', 'PRD-004', 'Marcadores', 'Caja de 12 permanentes', 30, 'Caja', 'DISP', '', ''],
    ]
    
    for row_idx, ejemplo in enumerate(ejemplos, 3):
        for col_idx, value in enumerate(ejemplo, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            if col_idx in [1, 2, 7]:  # Códigos y estado
                cell.font = Font(bold=True)
    
    # Ajustar anchos
    anchos = [15, 15, 20, 30, 12, 12, 12, 12, 30]
    for idx, ancho in enumerate(anchos, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = ancho
    
    # Respuesta
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=plantilla_productos.xlsx'
    wb.save(response)
    
    return response
import os
import json
import mimetypes
from datetime import datetime
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.utils import timezone
from django.http import FileResponse, Http404, HttpResponse
from django.views.decorators.clickjacking import xframe_options_exempt
from django.views.decorators.cache import cache_control
from django.views.decorators.csrf import csrf_exempt
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from .models import (
    ProductoAlmacen, 
    Unidad, 
    PedidoCompra, 
    ItemPedido, 
    Cotizacion,
    MovimientoInventario
)


# ========== VISTA PRINCIPAL: LISTA DE PEDIDOS ==========
def PedidosCompra(request):
    pedidos = PedidoCompra.objects.all()
    context = {
        'pedidos': pedidos,
        'total_pedidos': pedidos.count()
    }
    return render(request, 'almacenes/almgeneral/PedidosCompra.html', context)


# ========== CREAR NUEVO PEDIDO (WIZARD) ==========
def CrearPedidoCompra(request):
    if request.method == 'POST':
        nombre = request.POST.get('nombre')
        descripcion = request.POST.get('descripcion', '')
        archivo = request.FILES.get('archivo')  # Ahora es opcional
        productos_json = request.POST.get('productos_json')
        
        if not nombre:
            messages.error(request, 'El nombre del pedido es obligatorio')
            return redirect('CrearPedidoCompra')
        
        if not productos_json:
            messages.error(request, 'Debes seleccionar al menos un producto')
            return redirect('CrearPedidoCompra')
        
        try:
            # Parsear productos
            productos = json.loads(productos_json)
            
            if len(productos) == 0:
                messages.error(request, 'Debes seleccionar al menos un producto')
                return redirect('CrearPedidoCompra')
            
            # Validar archivo si se subió
            if archivo:
                extension = archivo.name.split('.')[-1].lower()
                tipos_permitidos = ['pdf', 'doc', 'docx']
                
                if extension not in tipos_permitidos:
                    messages.error(request, 'Solo se permiten archivos PDF o Word (.pdf, .doc, .docx)')
                    return redirect('CrearPedidoCompra')
            
            # Crear pedido
            pedido = PedidoCompra.objects.create(
                nombre=nombre,
                descripcion=descripcion,
                archivo=archivo if archivo else None
            )
            
            # Crear items del pedido
            for producto_data in productos:
                producto = ProductoAlmacen.objects.get(id_producto=producto_data['id'])
                ItemPedido.objects.create(
                    pedido=pedido,
                    producto=producto,
                    cantidad_solicitada=producto_data['cantidad'],
                    precio_unitario=producto_data['precio'],
                    observaciones=''
                )
            
            messages.success(request, f'Pedido "{pedido.nombre}" creado exitosamente con {len(productos)} productos')
            return redirect('PedidosCompra')
            
        except Exception as e:
            messages.error(request, f'Error al crear el pedido: {str(e)}')
            return redirect('CrearPedidoCompra')
    
    # GET request
    productos = ProductoAlmacen.objects.all().order_by('nombre')
    context = {
        'productos': productos,
    }
    return render(request, 'almacenes/almgeneral/CrearPedidoCompra.html', context)


# ========== GENERAR PDF DEL PEDIDO ==========
def GenerarPDFPedido(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.POST.get('data'))
            nombre = data.get('nombre', 'Pedido de Compra')
            descripcion = data.get('descripcion', '')
            productos = data.get('productos', [])
            
            # Crear respuesta HTTP
            response = HttpResponse(content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="Pedido_{nombre.replace(" ", "_")}.pdf"'
            
            # Crear PDF
            doc = SimpleDocTemplate(response, pagesize=letter)
            elements = []
            styles = getSampleStyleSheet()
            
            # Estilo personalizado para el título
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=20,
                textColor=colors.HexColor('#148129'),
                spaceAfter=30,
                alignment=TA_CENTER
            )
            
            # Título
            title = Paragraph(f"PEDIDO DE COMPRA", title_style)
            elements.append(title)
            
            # Información del pedido
            info_data = [
                ['Nombre del Pedido:', nombre],
                ['Fecha de Generación:', datetime.now().strftime('%d/%m/%Y %H:%M')],
                ['Descripción:', descripcion if descripcion else 'Sin descripción'],
            ]
            
            info_table = Table(info_data, colWidths=[2*inch, 4.5*inch])
            info_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f5f5f5')),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
                ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
                ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ]))
            
            elements.append(info_table)
            elements.append(Spacer(1, 20))
            
            # Título de tabla de productos
            subtitle = Paragraph("PRODUCTOS SOLICITADOS", styles['Heading2'])
            elements.append(subtitle)
            elements.append(Spacer(1, 10))
            
            # Tabla de productos
            table_data = [
                ['N°', 'Código', 'Producto', 'Cantidad', 'Unidad', 'Precio Unit.', 'Subtotal']
            ]
            
            total_general = 0
            
            for i, producto in enumerate(productos, 1):
                subtotal = producto['cantidad'] * producto['precio']
                total_general += subtotal
                
                table_data.append([
                    str(i),
                    producto['codigo'],
                    producto['nombre'],
                    str(producto['cantidad']),
                    producto['unidad'],
                    f"S/. {producto['precio']:.2f}",
                    f"S/. {subtotal:.2f}"
                ])
            
            # Agregar fila de total
            table_data.append([
                '', '', '', '', '', 'TOTAL:', f"S/. {total_general:.2f}"
            ])
            
            products_table = Table(table_data, colWidths=[0.4*inch, 1*inch, 2.5*inch, 0.8*inch, 0.8*inch, 1*inch, 1*inch])
            products_table.setStyle(TableStyle([
                # Header
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#148129')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                
                # Body
                ('TEXTCOLOR', (0, 1), (-1, -2), colors.black),
                ('ALIGN', (0, 1), (0, -2), 'CENTER'),  # N°
                ('ALIGN', (1, 1), (1, -2), 'LEFT'),    # Código
                ('ALIGN', (2, 1), (2, -2), 'LEFT'),    # Producto
                ('ALIGN', (3, 1), (3, -2), 'CENTER'),  # Cantidad
                ('ALIGN', (4, 1), (4, -2), 'CENTER'),  # Unidad
                ('ALIGN', (5, 1), (-1, -2), 'RIGHT'),  # Precios
                ('FONTNAME', (0, 1), (-1, -2), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -2), 9),
                
                # Total row
                ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#f5f5f5')),
                ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, -1), (-1, -1), 11),
                ('ALIGN', (0, -1), (-2, -1), 'RIGHT'),
                ('ALIGN', (-1, -1), (-1, -1), 'RIGHT'),
                
                # Grid
                ('GRID', (0, 0), (-1, -1), 1, colors.grey),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ]))
            
            elements.append(products_table)
            elements.append(Spacer(1, 30))
            
            # Footer
            footer_text = f"Total de productos: {len(productos)} | Total general: S/. {total_general:.2f}"
            footer = Paragraph(footer_text, styles['Normal'])
            elements.append(footer)
            
            # Generar PDF
            doc.build(elements)
            
            return response
            
        except Exception as e:
            messages.error(request, f'Error al generar PDF: {str(e)}')
            return redirect('CrearPedidoCompra')
    
    return redirect('CrearPedidoCompra')


# ========== VER DETALLES DEL PEDIDO ==========
def DetallePedido(request, id_pedido):
    pedido = get_object_or_404(PedidoCompra, id_pedido=id_pedido)
    items = pedido.items.all()
    
    context = {
        'pedido': pedido,
        'items': items,
        'total_items': items.count(),
    }
    return render(request, 'almacenes/almgeneral/DetallePedido.html', context)


# ========== EDITAR PEDIDO ==========
def EditarPedido(request, id_pedido):
    pedido = get_object_or_404(PedidoCompra, id_pedido=id_pedido)
    
    if request.method == 'POST':
        nombre = request.POST.get('nombre')
        descripcion = request.POST.get('descripcion', '')
        archivo = request.FILES.get('archivo')
        
        if not nombre:
            messages.error(request, 'El nombre del pedido es obligatorio')
            return redirect('EditarPedido', id_pedido=id_pedido)
        
        pedido.nombre = nombre
        pedido.descripcion = descripcion
        
        if archivo:
            extension = archivo.name.split('.')[-1].lower()
            tipos_permitidos = ['pdf', 'doc', 'docx']
            
            if extension not in tipos_permitidos:
                messages.error(request, 'Solo se permiten archivos PDF o Word')
                return redirect('EditarPedido', id_pedido=id_pedido)
            
            pedido.archivo = archivo
        
        try:
            pedido.save()
            messages.success(request, f'Pedido "{pedido.nombre}" actualizado exitosamente')
            return redirect('PedidosCompra')
        except Exception as e:
            messages.error(request, f'Error al actualizar el pedido: {str(e)}')
            return redirect('EditarPedido', id_pedido=id_pedido)
    
    context = {
        'pedido': pedido,
    }
    return render(request, 'almacenes/almgeneral/EditarPedido.html', context)


# ========== ELIMINAR PEDIDO ==========
def EliminarPedido(request, id_pedido):
    pedido = get_object_or_404(PedidoCompra, id_pedido=id_pedido)
    
    if request.method == 'POST':
        nombre = pedido.nombre
        try:
            pedido.delete()
            messages.success(request, f'Pedido "{nombre}" eliminado exitosamente')
        except Exception as e:
            messages.error(request, f'Error al eliminar el pedido: {str(e)}')
        return redirect('PedidosCompra')
    
    context = {
        'pedido': pedido,
    }
    return render(request, 'almacenes/almgeneral/EliminarPedido.html', context)


# ========== VER COTIZACIONES DEL PEDIDO ==========
def CotizacionesPedido(request, id_pedido):
    pedido = get_object_or_404(PedidoCompra, id_pedido=id_pedido)
    cotizaciones = pedido.cotizaciones.all()
    
    context = {
        'pedido': pedido,
        'cotizaciones': cotizaciones,
        'total_cotizaciones': cotizaciones.count()
    }
    return render(request, 'almacenes/almgeneral/CotizacionesPedido.html', context)


# ========== AGREGAR COTIZACIÓN ==========
def AgregarCotizacion(request, id_pedido):
    pedido = get_object_or_404(PedidoCompra, id_pedido=id_pedido)
    
    if request.method == 'POST':
        proveedor = request.POST.get('proveedor')
        monto = request.POST.get('monto')
        descripcion = request.POST.get('descripcion', '')
        documento = request.FILES.get('documento')
        
        if not proveedor:
            messages.error(request, 'El nombre del proveedor es obligatorio')
            return redirect('AgregarCotizacion', id_pedido=id_pedido)
        
        if not monto:
            messages.error(request, 'El monto es obligatorio')
            return redirect('AgregarCotizacion', id_pedido=id_pedido)
        
        if not documento:
            messages.error(request, 'Debes subir el documento de cotización')
            return redirect('AgregarCotizacion', id_pedido=id_pedido)
        
        extension = documento.name.split('.')[-1].lower()
        tipos_permitidos = ['pdf', 'doc', 'docx']
        
        if extension not in tipos_permitidos:
            messages.error(request, 'Solo se permiten archivos PDF o Word')
            return redirect('AgregarCotizacion', id_pedido=id_pedido)
        
        try:
            cotizacion = Cotizacion.objects.create(
                pedido=pedido,
                proveedor=proveedor,
                monto=monto,
                descripcion=descripcion,
                documento=documento
            )
            messages.success(request, f'Cotización de "{proveedor}" agregada exitosamente')
            return redirect('CotizacionesPedido', id_pedido=id_pedido)
        except Exception as e:
            messages.error(request, f'Error al agregar cotización: {str(e)}')
            return redirect('AgregarCotizacion', id_pedido=id_pedido)
    
    context = {
        'pedido': pedido,
    }
    return render(request, 'almacenes/almgeneral/AgregarCotizacion.html', context)


# ========== SELECCIONAR COTIZACIÓN GANADORA ==========
def SeleccionarCotizacion(request, id_cotizacion):
    cotizacion = get_object_or_404(Cotizacion, id_cotizacion=id_cotizacion)
    pedido = cotizacion.pedido
    
    if request.method == 'POST':
        try:
            pedido.cotizaciones.exclude(id_cotizacion=id_cotizacion).update(estado='RECH')
            cotizacion.estado = 'SELEC'
            cotizacion.save()
            pedido.estado = 'COMP'
            pedido.save()
            
            messages.success(request, f'Cotización de "{cotizacion.proveedor}" seleccionada. Pedido marcado como COMPLETADO')
            return redirect('CotizacionesPedido', id_pedido=pedido.id_pedido)
        except Exception as e:
            messages.error(request, f'Error al seleccionar cotización: {str(e)}')
            return redirect('CotizacionesPedido', id_pedido=pedido.id_pedido)
    
    context = {
        'cotizacion': cotizacion,
        'pedido': pedido,
    }
    return render(request, 'almacenes/almgeneral/SeleccionarCotizacion.html', context)


# ========== MARCAR COMO ENTREGADO ==========
def MarcarEntregado(request, id_pedido):
    pedido = get_object_or_404(PedidoCompra, id_pedido=id_pedido)
    
    if pedido.estado != 'COMP':
        messages.error(request, 'Solo puedes marcar como entregado pedidos que están COMPLETADOS')
        return redirect('CotizacionesPedido', id_pedido=id_pedido)
    
    if request.method == 'POST':
        documento = request.FILES.get('documento_entrega')
        
        if not documento:
            messages.error(request, 'Debes subir el documento de entrega')
            return redirect('MarcarEntregado', id_pedido=id_pedido)
        
        extension = documento.name.split('.')[-1].lower()
        tipos_permitidos = ['pdf', 'doc', 'docx']
        
        if extension not in tipos_permitidos:
            messages.error(request, 'Solo se permiten archivos PDF o Word')
            return redirect('MarcarEntregado', id_pedido=id_pedido)
        
        try:
            pedido.documento_entrega = documento
            pedido.fecha_entrega = timezone.now()
            pedido.estado = 'ENTR'
            pedido.save()
            
            messages.success(request, f'Pedido "{pedido.nombre}" marcado como ENTREGADO')
            return redirect('DetallePedido', id_pedido=id_pedido)
        except Exception as e:
            messages.error(request, f'Error al marcar como entregado: {str(e)}')
            return redirect('MarcarEntregado', id_pedido=id_pedido)
    
    context = {
        'pedido': pedido,
    }
    return render(request, 'almacenes/almgeneral/MarcarEntregado.html', context)


# ========== VER DOCUMENTO ==========
@xframe_options_exempt
@cache_control(max_age=3600, public=True)
def ver_documento(request, cotizacion_id):
    """Vista para servir documentos sin servicios externos"""
    cotizacion = get_object_or_404(Cotizacion, id_cotizacion=cotizacion_id)
    
    if not cotizacion.documento:
        raise Http404("Documento no encontrado")
    
    file_path = cotizacion.documento.path
    
    if not os.path.exists(file_path):
        raise Http404("Archivo físico no encontrado")
    
    filename = os.path.basename(file_path)
    extension = filename.split('.')[-1].lower()
    
    mime_types = {
        'pdf': 'application/pdf',
        'doc': 'application/msword',
        'docx': 'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
    }
    
    content_type = mime_types.get(extension, 'application/octet-stream')
    
    try:
        response = FileResponse(open(file_path, 'rb'), content_type=content_type)
        response['Content-Disposition'] = f'inline; filename="{filename}"'
        response['X-Content-Type-Options'] = 'nosniff'
        response['Accept-Ranges'] = 'bytes'
        
        if extension == 'pdf':
            response['Content-Type'] = 'application/pdf'
        
        return response
    except Exception as e:
        raise Http404(f"Error: {str(e)}")


# ========== AGREGAR PRODUCTO (REUTILIZADO) ==========
def agregar_producto(request):
    if request.method == 'POST':
        try:
            unidad_nombre = request.POST.get('unidad')
            unidad = Unidad.objects.get(nombre=unidad_nombre)
            
            producto = ProductoAlmacen.objects.create(
                nombre=request.POST.get('nombre'),
                descripcion=request.POST.get('descripcion'),
                ubicacion_almacen=request.POST.get('ubicacion_almacen'),
                estante=request.POST.get('estante'),
                cantidad=request.POST.get('cantidad'),
                unidad=unidad,
                estado=request.POST.get('estado'),
                observaciones=request.POST.get('observaciones', '')
            )
            
            messages.success(request, f'Producto "{producto.nombre}" agregado exitosamente')
            return redirect('InventarioAG')
            
        except Exception as e:
            messages.error(request, f'Error al agregar producto: {str(e)}')
    
    unidades = Unidad.objects.all().order_by('nombre')
    context = {
        'unidades': unidades
    }
    return render(request, 'almacenes/almgeneral/agregar_producto.html', context)


# ========== CREAR UNIDAD (AJAX) ==========
@csrf_exempt
def crear_unidad(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            nombre = data.get('nombre')
            abreviatura = data.get('abreviatura', '')
            
            if not nombre:
                return JsonResponse({'success': False, 'error': 'El nombre es obligatorio'})
            
            if Unidad.objects.filter(nombre=nombre).exists():
                return JsonResponse({'success': False, 'error': 'Esta unidad ya existe'})
            
            unidad = Unidad.objects.create(
                nombre=nombre,
                abreviatura=abreviatura
            )
            
            return JsonResponse({
                'success': True,
                'unidad': {
                    'nombre': unidad.nombre,
                    'abreviatura': unidad.abreviatura
                }
            })
            
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Método no permitido'})

def api_ultimo_producto(request):
    """API para obtener el último producto creado"""
    try:
        ultimo_producto = ProductoAlmacen.objects.latest('id_producto')
        
        data = {
            'id_producto': ultimo_producto.id_producto,
            'codigo_producto': ultimo_producto.codigo_producto,
            'nombre': ultimo_producto.nombre,
            'cantidad': ultimo_producto.cantidad,
            'unidad': ultimo_producto.unidad.nombre if ultimo_producto.unidad else '',
            'descripcion': ultimo_producto.descripcion,
        }
        
        return JsonResponse(data)
        
    except ProductoAlmacen.DoesNotExist:
        return JsonResponse({'error': 'No hay productos'}, status=404)


# ===== AGREGAR ESTA RUTA A urls.py =====